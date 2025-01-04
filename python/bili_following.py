import requests
import time
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def get_following_list(vmid, cookie):
    """获取用户关注列表"""

    # 设置请求头来绕过反爬措施
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Referer': 'https://space.bilibili.com/',
        'Cookie': cookie
    }

    # 存储所有关注数据
    all_followings = []
    page = 1

    while True:
        # 构造请求参数
        params = {
            'vmid': vmid,
            'ps': 50,  # 每页数量
            'pn': page,
            'order_type': ''  # 按关注顺序排序
        }

        try:
            # 发起请求
            response = requests.get(
                'https://api.bilibili.com/x/relation/followings',
                params=params,
                headers=headers
            )

            # 检查响应状态
            data = response.json()
            if data['code'] != 0:
                print(f"Error: {data['message']}")
                break

            # 获取关注列表
            followings = data['data']['list']
            if not followings:  # 如果列表为空，说明已经到达末尾
                break

            # 处理每个关注的用户数据
            for following in followings:
                following_data = {
                    'UID': str(following['mid']),
                    '用户名': following['uname'],
                    '签名': following['sign'],
                    '关注时间': datetime.fromtimestamp(following['mtime']).strftime('%Y-%m-%d %H:%M:%S'),
                    '是否认证': '是' if following['official_verify']['type'] in [0, 1] else '否',
                    '认证信息': following['official_verify']['desc'],
                    '是否为大会员': '是' if following['vip']['vipStatus'] == 1 else '否'
                }
                all_followings.append(following_data)

            print(f"已获取第 {page} 页数据")

            # 随机延迟，避免请求过快
            time.sleep(random.uniform(1, 2))

            page += 1

        except Exception as e:
            print(f"Error occurred: {str(e)}")
            break

    return all_followings


def export_to_excel(data, output_file='bilibili_followings.xlsx'):
    """将数据导出为Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "关注列表"

    # 定义表头
    headers = ['UID', '用户名', '签名', '关注时间', '是否认证', '认证信息', '是否为大会员']
    
    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="00A1D6", end_color="00A1D6", fill_type="solid")  # B站蓝
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    for row, user in enumerate(data, 2):
        ws.cell(row=row, column=1, value=user['UID'])
        ws.cell(row=row, column=2, value=user['用户名'])
        ws.cell(row=row, column=3, value=user['签名'])
        ws.cell(row=row, column=4, value=user['关注时间'])
        ws.cell(row=row, column=5, value=user['是否认证'])
        ws.cell(row=row, column=6, value=user['认证信息'])
        ws.cell(row=row, column=7, value=user['是否为大会员'])

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
        ws.column_dimensions[column].width = adjusted_width

    # 保存文件
    wb.save(output_file)
    print(f"数据已导出到: {output_file}")


def main():
    # 获取用户输入
    vmid = input("请输入要查询的用户UID: ")
    cookie = input("请输入Cookie(SESSDATA): ")

    cookie = f'SESSDATA={cookie}'
    
    print("开始获取数据...")
    following_list = get_following_list(vmid, cookie)

    if following_list:
        # 生成包含时间戳的文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'bilibili_followings_{timestamp}.xlsx'

        export_to_excel(following_list, output_file)
        print(f"共导出 {len(following_list)} 条关注数据")
    else:
        print("未获取到数据")


if __name__ == "__main__":
    main()